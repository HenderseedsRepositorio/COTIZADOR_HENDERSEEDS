"""
Parse Campaña 24-25 JAM Excel into JSON for import into the campana app.
"""
import openpyxl
import json
import warnings
import glob
from datetime import datetime

warnings.filterwarnings('ignore')

# Find the file
pattern = "C:/Users/ALVARO/Documents/000-COTIZADOR GENERAL/APP Seguimiento campa*a/Campa*24-25*.xlsx"
files = glob.glob(pattern)
if not files:
    # Fallback
    files = glob.glob("C:/Users/ALVARO/Documents/000-COTIZADOR GENERAL/APP*/Campa*24-25*.xlsx")
if not files:
    raise FileNotFoundError(f"No file matched pattern")

filepath = files[0]
print(f"Reading: {filepath}")

wb = openpyxl.load_workbook(filepath, data_only=True)

# ─────────────────────────────────────────────
# 1. Parse LOTES from "1- LOTES" sheet
# ─────────────────────────────────────────────
ws_lotes = wb['1- LOTES']
lotes = []
lotes_map = {}  # code -> lote dict

print("\n=== Parsing LOTES (gruesa section) ===")
for r in range(4, ws_lotes.max_row + 1):
    code = ws_lotes.cell(r, 35).value  # col AE = index 34 => 1-based col 35
    if not code:
        continue
    code = str(code).strip()
    cultivo = ws_lotes.cell(r, 36).value or ""  # col AF (Cultivo Gruesa)
    cultivo = str(cultivo).strip()
    ha = ws_lotes.cell(r, 38).value  # col AH (Ha Gruesa)
    try:
        ha = float(ha) if ha else 0
    except (ValueError, TypeError):
        ha = 0

    lote = {
        "id": f"l_imported_{code}",
        "nombre": code,
        "hectareas": ha,
        "cultivo": cultivo
    }
    lotes.append(lote)
    lotes_map[code] = lote
    print(f"  Lote: {code} - {cultivo} - {ha} ha")

print(f"Total lotes: {len(lotes)}")

# ─────────────────────────────────────────────
# 2. Parse ORDENES from "2- ORDENES" sheet
# ─────────────────────────────────────────────
ws = wb['2- ORDENES']

# Find all order start rows
order_starts = []
for r in range(1, ws.max_row + 1):
    v = ws.cell(r, 1).value  # col A
    if v is not None:
        try:
            num = int(float(v))
            order_starts.append((r, num))
        except (ValueError, TypeError):
            pass

print(f"\n=== Parsing ORDENES ({len(order_starts)} orders found) ===")

def parse_date(val):
    """Convert date value to ISO string."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    # Try common formats
    for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"]:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return s

def cell(r, c):
    """Get cell value (1-based col)."""
    return ws.cell(r, c).value

ordenes = []
aplicadores_set = set()

for idx, (start_row, order_num) in enumerate(order_starts):
    # Determine end row
    if idx + 1 < len(order_starts):
        end_row = order_starts[idx + 1][0] - 1
    else:
        end_row = start_row + 20  # max block size

    # Parse metadata from the block
    tipo = ""
    cultivo = ""
    fecha_orden = ""
    fecha_aplicacion = ""
    aplicador = ""
    campana_val = ""
    ha_aplicadas = 0
    estado = "pendiente"

    # Collect lotes and products
    lote_rows = []  # rows that have lote data (col E non-empty)
    product_rows = []  # rows with "x" in col K

    for r in range(start_row, end_row + 1):
        col_b = cell(r, 2)  # col B (1-based col 2)
        col_c = cell(r, 3)  # col C

        # Metadata rows
        if col_b:
            b_str = str(col_b).strip()
            if "Aplicaci" in b_str:  # "Aplicación"
                tipo = str(col_c).strip() if col_c else ""
            elif b_str == "Cultivo":
                cultivo = str(col_c).strip() if col_c else ""
            elif "Fecha orden" in b_str:
                fecha_orden = parse_date(col_c)
            elif "Mano de obra" in b_str:
                aplicador = str(col_c).strip().lower() if col_c else ""
            elif "Campa" in b_str:
                campana_val = str(col_c).strip() if col_c else ""
            elif "Ha Aplicadas" in b_str:
                try:
                    ha_aplicadas = float(col_c) if col_c else 0
                except (ValueError, TypeError):
                    ha_aplicadas = 0

        # Check for lote data in cols E, F, G, H (1-based: 5, 6, 7, 8)
        lote_code = cell(r, 5)
        establec_raw = cell(r, 6)
        ha_raw = cell(r, 7)
        # A lote row has: lote code in E, OR establecimiento in F with ha in G
        has_lote_code = lote_code and str(lote_code).strip() not in ("Lotes", "")
        has_establec_data = (not has_lote_code and establec_raw and
                            str(establec_raw).strip() not in ("Establec.", "") and
                            ha_raw is not None)
        if has_lote_code or has_establec_data:
            # Use establecimiento as fallback lote code if lote_code missing
            if not has_lote_code:
                lote_code = str(establec_raw).strip()
            establec = str(establec_raw).strip() if establec_raw else ""
            ha_efect = ha_raw or 0
            fecha_app = cell(r, 8)
            estado_cell = cell(r, 23)  # col W = 1-based col 23

            try:
                ha_efect = float(ha_efect)
            except (ValueError, TypeError):
                ha_efect = 0

            if fecha_app and not fecha_aplicacion:
                fecha_aplicacion = parse_date(fecha_app)

            if estado_cell and str(estado_cell).strip() == "Aplicada":
                estado = "completada"

            lote_rows.append({
                "code": str(lote_code).strip(),
                "establecimiento": str(establec).strip(),
                "hectareas": ha_efect,
                "fecha": parse_date(fecha_app) if fecha_app else ""
            })

        # Product rows: col K (1-based col 11) = "x"
        col_k = cell(r, 11)
        if col_k and str(col_k).strip().lower() == "x":
            prod_name = cell(r, 12) or ""  # col L = 1-based col 12
            dosis = cell(r, 13) or 0       # col M = 1-based col 13
            total = cell(r, 15) or 0       # col O = 1-based col 15

            prod_name = str(prod_name).strip()
            if not prod_name or prod_name in ("Producto", ""):
                continue

            try:
                dosis = float(dosis)
            except (ValueError, TypeError):
                dosis = 0
            try:
                total = float(total)
            except (ValueError, TypeError):
                total = 0

            # Determine unit from product name
            unidad = "L"
            name_upper = prod_name.upper()
            if "KG" in name_upper or "WDG" in name_upper or "WG" in name_upper:
                unidad = "Kg"

            product_rows.append({
                "producto": prod_name,
                "dosis": round(dosis, 4),
                "total": round(total, 2),
                "unidad": unidad
            })

    # Use fecha_aplicacion or fall back to fecha_orden
    fecha = fecha_aplicacion or fecha_orden or ""

    if aplicador:
        aplicadores_set.add(aplicador)

    # Build lotes for aplicaciones
    app_lotes = []
    for lr in lote_rows:
        lote_code = lr["code"]
        # Try to find cultivo from lotes_map
        lote_cultivo = cultivo  # default from order
        if lote_code in lotes_map:
            lote_cultivo = lotes_map[lote_code].get("cultivo", cultivo)

        app_lotes.append({
            "loteId": f"l_imported_{lote_code}",
            "hectareas": lr["hectareas"],
            "cultivo": lote_cultivo
        })

    # Skip empty template orders (no lotes, no real products, no fecha)
    has_real_products = any(p["producto"] for p in product_rows)
    if not app_lotes and not has_real_products:
        print(f"  Order {order_num}: SKIPPED (empty template)")
        continue
    if not fecha and not app_lotes and ha_aplicadas == 0:
        print(f"  Order {order_num}: SKIPPED (no date, no lotes, 0 ha)")
        continue

    # Build products list
    productos = []
    for i, p in enumerate(product_rows):
        if p["producto"]:  # skip empty products
            productos.append({
                "id": i + 1,
                "producto": p["producto"],
                "dosis": p["dosis"],
                "precio": 0,
                "unidad": p["unidad"],
                "formulacion": "",
                "categoria": "Herbicidas",
                "iva_pct": 0
            })

    created = f"{fecha}T12:00:00.000Z" if fecha else ""

    orden = {
        "id": f"ot_imported_{order_num}",
        "aplicador": aplicador,
        "fecha": fecha,
        "tipo": tipo,
        "costoLabor": 0,
        "estado": estado,
        "notas": "",
        "aplicaciones": [{
            "lotes": app_lotes,
            "productos": productos
        }],
        "createdAt": created,
        "updatedAt": created,
        "campana": "2024/25"
    }

    ordenes.append(orden)

    prod_count = len(productos)
    lote_count = len(app_lotes)
    print(f"  Order {order_num}: {tipo} | {cultivo} | {aplicador} | {fecha} | {estado} | {lote_count} lotes | {prod_count} products")

# ─────────────────────────────────────────────
# 3. Build final JSON
# ─────────────────────────────────────────────
aplicadores = sorted(aplicadores_set)

result = {
    "lotes": lotes,
    "ordenes": ordenes,
    "aplicadores": aplicadores,
    "campana": "2024/25"
}

output_path = "C:/Users/ALVARO/Documents/000-COTIZADOR GENERAL/import-jam-24-25.json"
with open(output_path, "w", encoding="utf-8") as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

print(f"\n=== DONE ===")
print(f"Output: {output_path}")
print(f"Lotes: {len(lotes)}")
print(f"Ordenes: {len(ordenes)}")
print(f"Aplicadores: {aplicadores}")
print(f"Total products across all orders: {sum(len(o['aplicaciones'][0]['productos']) for o in ordenes)}")
